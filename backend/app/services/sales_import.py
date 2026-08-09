"""Parse sale data (Excel/CSV/PDF/free text) into candidate leads.

Best-effort parsing that maps common column headers and line shapes onto the
Lead model. It is deliberately lenient: rows that cannot be understood are
skipped rather than failing the whole import.
"""

import csv
import io
import re
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

COMPANY_HEADERS = {
    "company",
    "company name",
    "title",
    "business",
    "business name",
    "client",
    "client name",
    "account",
    "organization",
    "organisation",
    "org",
    "prospect",
    "lead",
    "name",
    "vendor",
    "brand",
}
CONTACT_HEADERS = {
    "contact",
    "contact name",
    "contact person",
    "person",
    "full name",
    "representative",
    "owner",
    "sales rep",
    "rep",
    "primary contact",
}
PHONE_HEADERS = {
    "phone",
    "phone number",
    "mobile",
    "mobile number",
    "tel",
    "telephone",
    "cell",
    "cell number",
    "contact no",
    "contact number",
}
EMAIL_HEADERS = {"email", "email address", "e-mail", "e-mail address", "mail", "mail id"}
VALUE_HEADERS = {
    "value",
    "lead value",
    "deal value",
    "amount",
    "revenue",
    "price",
    "estimated value",
    "annual value",
    "project value",
    "budget",
    "worth",
}
STATUS_HEADERS = {"status", "stage", "lead status", "pipeline stage", "phase"}
WEBSITE_HEADERS = {
    "website",
    "url",
    "web",
    "site",
    "site url",
    "domain",
    "website url",
    "web address",
    "link",
}
CATEGORY_HEADERS = {
    "category",
    "categoryname",
    "category name",
    "industry",
    "business type",
    "business category",
    "type",
    "sector",
    "segment",
}
ADDRESS_HEADERS = {
    "address",
    "location",
    "street",
    "street address",
    "city",
    "area",
    "locality",
}
NOTE_HEADERS = {
    "notes",
    "stage note",
    "notes/notes",
    "stage_note",
    "note",
    "remark",
    "remarks",
    "comments",
    "comment",
    "details",
}

_STATUS_SYNONYMS = {
    "new": {"new", "n", "cold", "fresh", "lead", "inbound"},
    "contacted": {"contacted", "contact", "reached out", "in touch", "touched", "warm"},
    "closing": {"closing", "close", "negotiation", "proposal", "sent proposal", "hot"},
    "closed": {"closed", "won", "signed", "done", "deal", "customer", "client"},
}

_EMAIL_RE = re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}")
_MONEY_RE = re.compile(r"(?:\$|€|£|₹)?\s?([\d][\d,]*(?:\.\d{1,2})?)\s*(?:k|K|usd|USD)?")
_PHONE_RE = re.compile(r"(\+?\d[\d\s\-\.]{7,}\d)")


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_header(h: Any) -> str:
    return _clean(h).strip().lower().replace("_", " ").replace("\u00a0", " ")


def _map_row(headers: list[str], row: list[Any]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for i, raw in enumerate(row):
        if i >= len(headers) or _clean(raw) == "":
            continue
        key = _normalize_header(headers[i])
        if key in COMPANY_HEADERS and not mapping.get("company"):
            mapping["company"] = _clean(raw)
        elif key in CONTACT_HEADERS and not mapping.get("contact"):
            mapping["contact"] = _clean(raw)
        elif key in PHONE_HEADERS and not mapping.get("phone"):
            mapping["phone"] = _clean(raw)
        elif key in EMAIL_HEADERS and not mapping.get("email"):
            mapping["email"] = _clean(raw)
        elif key in VALUE_HEADERS and not mapping.get("value"):
            mapping["value"] = _clean(raw)
        elif key in STATUS_HEADERS and not mapping.get("status"):
            mapping["status"] = _clean(raw)
        elif key in WEBSITE_HEADERS and not mapping.get("website"):
            mapping["website"] = _clean(raw)
        elif key in CATEGORY_HEADERS and not mapping.get("category"):
            mapping["category"] = _clean(raw)
        elif key in ADDRESS_HEADERS and not mapping.get("address"):
            mapping["address"] = _clean(raw)
        elif key in NOTE_HEADERS and not mapping.get("note"):
            mapping["note"] = _clean(raw)
    return mapping


def _parse_value(raw: str) -> float | None:
    text = _clean(raw)
    if not text:
        return None
    match = _MONEY_RE.search(text)
    if not match:
        return None
    number = match.group(1).replace(",", "")
    try:
        return float(number)
    except ValueError:
        return None


def _parse_status(raw: str) -> str | None:
    text = _clean(raw).strip().lower()
    if not text:
        return None
    for status, words in _STATUS_SYNONYMS.items():
        if text in words:
            return status
        if any(word in text for word in words if len(word) >= 3):
            return status
    return None


def _email_from_part(raw: str) -> str | None:
    match = _EMAIL_RE.search(_clean(raw))
    return match.group(0) if match else None


def _phone_from_part(raw: str) -> str | None:
    text = _clean(raw)
    match = _PHONE_RE.search(text)
    if not match:
        return None
    digits = re.sub(r"[^\d+]", "", match.group(0))
    if len(digits) < 8 or digits.count("+") > 1 or digits.endswith("+"):
        return None
    return digits[:15]


def _stage_note_from_part(raw: str) -> str | None:
    text = _clean(raw).strip(" -|:,")
    return text[:400] if text else None


def build_lead(
    company: str = "",
    contact: str = "",
    email: str = "",
    value: str = "",
    status: str = "",
    note: str = "",
    phone: str = "",
    website: str = "",
    category: str = "",
    address: str = "",
) -> dict | None:
    company = _clean(company).strip(" :|,")
    if not company:
        company = None
    mail = _email_from_part(email) or _email_from_part(contact)
    parsed_value = _parse_value(value)
    parsed_status = _parse_status(status)
    stage_note = _stage_note_from_part(note) or None
    if company is None and mail is None:
        return None
    return {
        "company": company or (mail.split("@")[1] if mail else "Unknown"),
        "contact": _clean(contact).strip(" :|,") or None,
        "phone": _clean(phone).strip() or None,
        "email": mail,
        "category": _clean(category).strip() or None,
        "address": _clean(address).strip() or None,
        "website": _clean(website).strip() or None,
        "value": parsed_value or 0.0,
        "status": parsed_status or "new",
        "stage_note": stage_note,
    }


def parse_rows(headers: list[str], rows: list[list[Any]]) -> tuple[list[dict], int]:
    leads: list[dict] = []
    for row in rows:
        if not any(_clean(c) for c in row):
            continue
        mapped = _map_row(headers, row)
        lead = build_lead(
            mapped.get("company", ""),
            mapped.get("contact", ""),
            mapped.get("email", ""),
            mapped.get("value", ""),
            mapped.get("status", ""),
            mapped.get("note", ""),
            mapped.get("phone", ""),
            mapped.get("website", ""),
            mapped.get("category", ""),
            mapped.get("address", ""),
        )
        if lead:
            leads.append(lead)
    return leads, 0


def parse_tsv_text(text: str) -> tuple[list[dict], int]:
    """Parse tab/comma/pipe separated text or one-line-per-lead notes."""
    cleaned = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not cleaned:
        return [], 0

    first_line = cleaned.split("\n", 1)[0] if cleaned else ""
    if "\t" in first_line:
        delimiter = "\t"
    elif "|" in first_line:
        delimiter = "|"
    elif "," in first_line:
        delimiter = ","
    else:
        return parse_freeform([l for l in cleaned.split("\n") if l.strip()])

    try:
        reader = csv.reader(io.StringIO(cleaned), delimiter=delimiter)
        table = [row for row in reader if any(c.strip() for c in row)]
    except csv.Error:
        return parse_freeform([l for l in cleaned.split("\n") if l.strip()])
    if not table:
        return [], 0

    headers = [h.strip().lower() for h in table[0]]
    known = (
        COMPANY_HEADERS | CONTACT_HEADERS | PHONE_HEADERS | EMAIL_HEADERS
        | VALUE_HEADERS | STATUS_HEADERS | WEBSITE_HEADERS | CATEGORY_HEADERS
        | ADDRESS_HEADERS | NOTE_HEADERS
    )
    if any(_normalize_header(h) in known for h in headers):
        return parse_rows(table[0], table[1:])
    return parse_freeform([l for l in cleaned.split("\n") if l.strip()])


def parse_freeform(lines: list[str]) -> tuple[list[dict], int]:
    leads: list[dict] = []
    skipped = 0
    for line in lines:
        line = _clean(line).strip(" \t-*•·")
        if not line:
            continue

        email = _email_from_part(line)
        value_raw = _parse_value(line)
        status = _parse_status(line)
        phone = _phone_from_part(line)
        note_match = re.search(r"[—\-–]\s*([^—\-–]+)$", line)
        stage_note = _stage_note_from_part(note_match.group(1)) if note_match else None

        remainder = line
        if email:
            remainder = remainder.replace(email, " ")
        if phone:
            remainder = remainder.replace(phone, " ")
        remainder = re.sub(r"\$?[\d][\d,]*(?:\.\d{1,2})?", "", remainder)
        if stage_note and email and email in stage_note:
            stage_note = None
        if stage_note:
            idx = remainder.rfind(" - ")
            remainder = remainder[:idx] if idx != -1 else remainder
        tokens = [t for t in re.split(r"[\t|,;:]+|\s{1,}", remainder.strip()) if t]
        if not tokens:
            skipped += 1
            continue
        company = tokens[0]
        contact = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        contact = _clean(contact).strip(" -—–")
        leads.append(
            {
                "company": company,
                "contact": contact or None,
                "email": email,
                "phone": phone,
                "value": value_raw or 0.0,
                "status": status or "new",
                "stage_note": stage_note,
            }
        )
    return leads, skipped


def parse_xlsx_bytes(data: bytes) -> tuple[list[dict], int]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return [], 0
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter, None)
    except StopIteration:
        header_row = None
    if header_row is None:
        return [], 0
    rows = list(rows_iter)
    leads, skipped = parse_rows(header_row, rows)
    return leads, skipped


def parse_csv_bytes(data: bytes) -> tuple[list[dict], int]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = data.decode("latin-1")
        except UnicodeDecodeError:
            return [], 0
    return parse_tsv_text(text)


def parse_pdf_bytes(data: bytes) -> tuple[list[dict], int]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return [], 0
    if not text.strip():
        return [], 0
    return parse_tsv_text(text)


def parse_text_input(text: str) -> tuple[list[dict], int]:
    return parse_tsv_text(text)


def parse_file(filename: str, data: bytes) -> tuple[list[dict], int]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    if name.endswith(".csv") or name.endswith(".tsv"):
        return parse_csv_bytes(data)
    if name.endswith(".pdf"):
        return parse_pdf_bytes(data)
    return [], 0


def dedupe_leads(leads: list[dict]) -> list[dict]:
    seen: set[str] = set()
    out: list[dict] = []
    by_email: dict[str, dict] = {}
    for lead in leads:
        email = (lead.get("email") or "").strip().lower()
        if email and email in by_email:
            by_email[email].update({k: v for k, v in lead.items() if v})
            continue
        if email:
            by_email[email] = lead
            seen.add(email)
        else:
            key = (lead.get("company") or "").strip().lower()
            if key and key not in seen:
                seen.add(key)
                out.append(lead)
            elif not key:
                out.append(lead)
    out.extend(by_email.values())
    return out


def timestamp_token() -> str:
    return datetime.utcnow().strftime("%Y%m%d%H%M%S")